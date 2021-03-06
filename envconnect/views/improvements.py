# Copyright (c) 2018, DjaoDjin inc.
# see LICENSE.
from __future__ import unicode_literals

import csv, datetime, logging, io

from django.core.urlresolvers import reverse
from django.http import HttpResponse
from django.views.generic.list import ListView
from django.utils import six
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.fills import FILL_SOLID
from pages.models import PageElement
from survey.models import Answer, Question
from extended_templates.backends.pdf import PdfTemplateResponse

from ..mixins import ImprovementQuerySetMixin
from ..models import Consumption
from .benchmark import PrintableChartsMixin
from .assessments import AssessmentView


LOGGER = logging.getLogger(__name__)


class ImprovementView(ImprovementQuerySetMixin, AssessmentView):

    template_name = 'envconnect/improve.html'
    breadcrumb_url = 'improve'

    @staticmethod
    def get_scorecard_path(path):
        parts = path.split('/')
        for idx, part in enumerate(parts):
            if part.startswith('sustainability-'):
                return '/'.join(parts[:idx + 1])
        return path

    def get_breadcrumb_url(self, path):
        organization = self.kwargs.get('organization', None)
        if organization:
            return reverse(
                'envconnect_improve_organization', args=(organization, path))
        return super(ImprovementView, self).get_breadcrumb_url(path)

    def get_report_tree(self):
        self.get_or_create_improve_sample()
        root = super(ImprovementView, self).get_report_tree()
        if root:
            self.decorate_with_breadcrumbs(root)
        return root

    def get_context_data(self, **kwargs):
        context = super(ImprovementView, self).get_context_data(**kwargs)
        organization = context['organization']
        self.update_context_urls(context, {
            'api_account_benchmark': reverse('api_benchmark',
                args=(organization, self.get_scorecard_path(
                    self.kwargs.get('path'))))})
        return context


class ImprovementSpreadsheetView(ImprovementQuerySetMixin, ListView):

    basename = 'improvements'
    headings = ['Practice', 'Implementation rate', 'Implemented by you?',
        'Opportunity score', 'Value']
    value_headings = ['', '', '', '',
        'Environmental', 'Ops/maintenance', 'Financial',
        'Implementation ease', 'AVERAGE VALUE']
    indent_step = '    '

    def insert_path(self, tree, parts=None):
        if not parts:
            return tree
        if not parts[0] in tree:
            tree[parts[0]] = {}
        return self.insert_path(tree[parts[0]], parts[1:])

    def write_tree(self, root, indent=''):
        for element in sorted(
                list(root.keys()), key=lambda node: (node.tag, node.pk)):
            # XXX sort won't exactly match the web presentation
            # which uses RelationShip order
            # (see ``BreadcrumbMixin._build_tree``).
            nodes = root[element]
            if 'opportunity' in nodes:
                # We reached a leaf
                self.writerow([
                    indent + element.title,
                    nodes['rate'],
                    'Yes',
                    nodes['opportunity'],
                    nodes['environmental_value'],
                    nodes['business_value'],
                    nodes['implementation_ease'],
                    nodes['profitability'],
                    nodes['avg_value']
                ], leaf=True)
            else:
                self.writerow([indent + element.title])
                self.write_tree(nodes, indent=indent + self.indent_step)

    def get(self, request, *args, **kwargs): #pylint: disable=unused-argument
        opportunities = {}
        for consumption in Consumption.objects.with_opportunity(
                population=Consumption.objects.get_active_by_accounts(
                    excludes=self._get_filter_out_testing())):
            opportunities[consumption.pk] = consumption

        self.root = {}
        for improvement in self.get_queryset():
            consumption = improvement.question
            _, parts = self.get_breadcrumbs(consumption.path)
            leaf = self.insert_path(self.root, [part[0] for part in parts])
            details = opportunities[consumption.pk]
            page_element = parts[-1][0]
            leaf.update({
                'path': consumption.path,
                'rate': consumption.rate,
                'opportunity': details.opportunity,
                'environmental_value': consumption.environmental_value,
                'business_value': consumption.business_value,
                'implementation_ease': consumption.implementation_ease,
                'profitability': consumption.profitability,
                'avg_value': consumption.avg_value,
                'text': page_element.text
            })

        self.create_writer(self.value_headings, title="Improvement Plan")
        self.writerow(
            ["Improvement Plan  -  Environmental practices"], leaf=True)
        self.writerow(self.get_headings(), leaf=True)
        self.writerow(self.value_headings, leaf=True)
        self.write_tree(self.root)
        resp = HttpResponse(self.flush_writer(), content_type=self.content_type)
        resp['Content-Disposition'] = \
            'attachment; filename="{}"'.format(self.get_filename())
        return resp

    def get_headings(self):
        return self.headings

    def get_filename(self):
        return datetime.datetime.now().strftime(self.basename + '-%Y%m%d.csv')


class ImprovementCSVView(ImprovementSpreadsheetView):

    content_type = 'text/csv'

    def writerow(self, row, leaf=False):
        #pylint:disable=unused-argument
        self.csv_writer.writerow([
            rec.encode('utf-8') if six.PY2 else rec
            for rec in row])

    def create_writer(self, headings, title=None):
        #pylint:disable=unused-argument
        self.content = io.StringIO()
        self.csv_writer = csv.writer(self.content)

    def flush_writer(self):
        self.content.seek(0)
        return self.content

    def get_filename(self):
        return datetime.datetime.now().strftime(self.basename + '-%Y%m%d.csv')


class ImprovementXLSXView(PrintableChartsMixin, ImprovementSpreadsheetView):

    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    valueFills = [
        PatternFill(fill_type=FILL_SOLID, fgColor='FF9CD76B'), # green-level-0
        PatternFill(fill_type=FILL_SOLID, fgColor='FF69B02B'), # green-level-1
        PatternFill(fill_type=FILL_SOLID, fgColor='FF007C3F'), # green-level-2
        PatternFill(fill_type=FILL_SOLID, fgColor='FFFFD700'), # green-level-3
    ]

    def create_writer(self, headings, title=None):
        col_scale = 11.9742857142857
        self.wbook = Workbook()
        self.wsheet = self.wbook.active
        if title:
            self.wsheet.title = title
        self.wsheet.row_dimensions[1].height = 0.36 * (6 * col_scale)
        self.wsheet.column_dimensions['A'].width = 6.56 * col_scale
        for col_num in range(0, len(headings)):
            self.wsheet.column_dimensions[chr(ord('E') + col_num)].width \
                = 0.99 * col_scale
        self.heading_font = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF0071BB')
        self.heading_alignment = Alignment(wrap_text=True)
        self.border = Border(
            left=Side(border_style=BORDER_THIN, color='FF000000'),
            right=Side(border_style=BORDER_THIN, color='FF000000'),
            top=Side(border_style=BORDER_THIN, color='FF000000'),
            bottom=Side(border_style=BORDER_THIN, color='FF000000'))
        self.text_center = Alignment(horizontal='center')

    def flush_writer(self):
        #pylint:disable=protected-access
        bold_font = Font(
            name='Calibri', size=11, bold=True, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        self.wsheet.append([])
        self.wsheet.append(["Implementation rate"])
        self.wsheet.row_dimensions[self.wsheet._current_row].font = bold_font
        self.wsheet.append(["    Percentage of peer"\
            " respondents that have implemented a best practice."])
        self.wsheet.append(["Implemented by you?"])
        self.wsheet.row_dimensions[self.wsheet._current_row].font = bold_font
        self.wsheet.append(["    Extent to which you"\
            " indicated the practice is implemented across"\
            " activities/services/products/offices/facilities to which"\
            " it could apply. 3 ticks = All, 2 ticks = More than 50%,"\
            " 1 tick = Less than 50%, X = Not implemented or Not applicable."])
        self.wsheet.append(["Opportunity score"])
        self.wsheet.row_dimensions[self.wsheet._current_row].font = bold_font
        self.wsheet.append(["    Percentage points by which"\
            " your score could increase if this best practice is implemented"\
            " to full extent. See FAQs for scoring methodology"\
            " and calculations."])
        alignment = Alignment(
            horizontal='center', vertical='center',
            text_rotation=0, wrap_text=False,
            shrink_to_fit=False, indent=0)
        title_fill = PatternFill(fill_type=FILL_SOLID, fgColor='FFDDD9C5')
        subtitle_fill = PatternFill(fill_type=FILL_SOLID, fgColor='FFEEECE2')
        subtitle_font = Font(
            name='Calibri', size=10, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        subtitle_alignment = Alignment(
            horizontal='center', vertical='center',
            text_rotation=0, wrap_text=True,
            shrink_to_fit=True, indent=0)
        self.wsheet.merge_cells('A1:I1')
        row = self.wsheet.row_dimensions[1]
        row.fill = title_fill
        row.font = Font(
            name='Calibri', size=20, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        row.alignment = alignment
        row.border = self.border
        self.wsheet.merge_cells('E2:I2')
        self.wsheet.merge_cells('A2:A3')
        self.wsheet.merge_cells('B2:B3')
        self.wsheet.merge_cells('C2:C3')
        self.wsheet.merge_cells('D2:D3')
        row = self.wsheet.row_dimensions[2]
        row.fill = subtitle_fill
        row.font = subtitle_font
        row.alignment = subtitle_alignment
        row.border = self.border
        row = self.wsheet.row_dimensions[3]
        row.fill = subtitle_fill
        row.font = subtitle_font
        row.alignment = subtitle_alignment
        row.border = self.border

        # Create "Impact of Improvement Plan" worksheet.
        wsheet = self.wbook.create_sheet(title="Impact of Improvement Plan")
        # XXX get data from actual score.
        chart = {
            'slug': 'totals',
            'highest_normalized_score': 100,
            'avg_normalized_score': 50,
            'normalized_score': 64,
        }
        self.generate_printable_html([chart])
        image_path = chart['image']
        if image_path.startswith('file://'):
            image_path = image_path[7:]
        img = Image(image_path)
        img.anchor(wsheet.cell('A1'))
        wsheet.add_image(img)

        # Write out the Excel file.
        content = io.BytesIO()
        self.wbook.save(content)
        content.seek(0)
        return content

    def get_filename(self):
        return datetime.datetime.now().strftime(self.basename + '-%Y%m%d.xlsx')

    def get_value_fill(self, val):
        if val < len(self.valueFills):
            return self.valueFills[val]
        return self.valueFills[-1]

    def write_best_practice_content(self, root, indent=''):
        for element in sorted(
                list(root.keys()), key=lambda node: (node.tag, node.pk)):
            # XXX sort won't exactly match the web presentation
            # which uses RelationShip order
            # (see ``BreadcrumbMixin._build_tree``).
            nodes = root[element]
            if 'text' in nodes:
                # We reached a leaf
                title = element.title.replace('*', '').replace('/', '')
                if len(title) > 31:
                    title = title[0:28] + '...'
                wsheet = self.wbook.create_sheet(title=title)
                wsheet['A1'] = nodes['text']
            else:
                self.write_best_practice_content(
                    nodes, indent=indent + self.indent_step)

    def writerow(self, row, leaf=False):
        #pylint:disable=protected-access
        self.wsheet.append(row)
        if leaf:
            if len(row) > 8:
                for row_cells in self.wsheet.iter_rows(
                        min_row=self.wsheet._current_row,
                        max_row=self.wsheet._current_row):
                    for cell_idx in range(4, 9):
                        # environmental_value, business_value,
                        # implementation_ease, profitability,
                        # avg_value
                        try:
                            row_cells[cell_idx].border = self.border
                            row_cells[cell_idx].fill = self.get_value_fill(
                                int(row_cells[cell_idx].value))
                            row_cells[cell_idx].value = None
                        except ValueError:
                            # might be the header
                            pass
                    row_cells[0].alignment = self.heading_alignment
                    row_cells[1].style = 'Percent'         # Implementation rate
                    row_cells[2].alignment = self.text_center # Imp. by you?
        else:
            for row_cells in self.wsheet.iter_rows(
                    min_row=self.wsheet._current_row,
                    max_row=self.wsheet._current_row):
                row_cells[0].font = self.heading_font
                row_cells[0].alignment = self.heading_alignment


class ReportPDFView(ImprovementQuerySetMixin, ListView):

    model = Question
    http_method_names = ['get']
    template_name = 'envconnect/best_practice_pdf.html'

    def get_queryset(self):
        return self.model.objects.filter(survey=self.sample).exclude(
            answer__text=Consumption.NOT_APPLICABLE)

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        report_items = []

        for question in self.object_list:
            report_items += [
                {
                    'answer': Answer.objects.get(question_id=question.id),
                    'consumption': Consumption.objects.get(id=question.text)
                }
            ]
        context = self.get_context_data(**kwargs)
        industry = PageElement.objects.get(slug=self.kwargs.get('industry'))
        context.update({'industry':industry})
        context.update({'report_items':report_items})
        return PdfTemplateResponse(request, self.template_name, context)
